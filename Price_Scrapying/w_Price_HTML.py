# -*- coding: utf-8 -*-
import datetime
import sqlite3 as lite
import array
import time
import re
#Scrape Data


SHtime=datetime.datetime.now()
print SHtime.minute
print SHtime.hour
#fetch last day (not today)
today = datetime.date.today( )
ytd = today - datetime.timedelta(days=1)
ymm =  '{:02d}'.format(ytd.month)
ydd =  '{:02d}'.format(ytd.day)
divYdate = ("%s-%s-%s"%(ytd.year,ymm,ydd))
print divYdate
print today

#fetch database n write to the html
#SQLite
con = lite.connect('daily_price.db')
with con:
    cur=con.cursor()
    getCount1=0
    while getCount1 != 1:
        sql_action='SELECT * FROM daily WHERE Record_Date=?'
        parameters = [divYdate]
        for row1 in con.execute(sql_action,parameters):
            print "find", row1[1], ytd
            #back a day!!! for next while loop
            ytd2 =ytd- datetime.timedelta(days=1)
            ymm2 =  '{:02d}'.format(ytd2.month)
            ydd2 =  '{:02d}'.format(ytd2.day)
            divYdate2 = ("%s-%s-%s"%(ytd2.year,ymm2,ydd2))
            getCount1 =1
            break
        else:
            ytd = ytd - datetime.timedelta(days=1)
            ymm =  '{:02d}'.format(ytd.month)
            ydd =  '{:02d}'.format(ytd.day)
            divYdate = ("%s-%s-%s"%(ytd.year,ymm,ydd))
            print ("NOT FOUND%s"%(divYdate))
    getCount2=0
    print divYdate2, "HAHA"
    while getCount2 != 1:
        sql_action='SELECT * FROM daily WHERE Record_Date=?'
        parameters = [divYdate2]
        for row2 in con.execute(sql_action,parameters):
            print "find", row2[1]
            getCount2 = 1
            break
        else:
            ytd2 = ytd2 - datetime.timedelta(days=1)
            ymm =  '{:02d}'.format(ytd2.month)
            ydd =  '{:02d}'.format(ytd2.day)
            divYdate2 = ("%s-%s-%s"%(ytd2.year,ymm,ydd))
            print ("NOT FOUND%s"%(divYdate2))


print row1
print row2
date1=row1[1].encode("utf-8")
date2=row2[1].encode("utf-8")
#write data in html file
f = open('price.html','w')

message = """<html>
<head><meta charset="UTF-8"></head>
<body style="font-family:'新細明體','Ariel'; font-size:'21px'" ><b>
<p style="color:blue">各位長官好:</p>
<p style="color:blue">LME國際金屬行情CASH BUYER價格(USD/TON):</p>"""
message1="""<p style="color:#FF3399">
<a>%s&nbsp&nbsp&nbsp 銅</a><a>%s,&nbsp&nbsp鋁%s,&nbsp&nbsp鎳%s,&nbsp&nbsp鋅%s</a></p>
<p style="color:#FF3399">
<a>%s&nbsp&nbsp&nbsp 銅</a><a>%s,&nbsp&nbsp鋁%s,&nbsp&nbsp鎳%s,&nbsp&nbsp鋅%s</a></p>
<p style="color:blue">滬銅行情價格(CNY/TON):</p>
<p style="color:#FF3399">%s&nbsp&nbsp%s</p>
<p style="color:#FF3399">%s&nbsp&nbsp%s</p>
"""%(date1,int(row1[2]),int(row1[3]),int(row1[4]),int(row1[5]),date2,int(row2[2]),int(row2[3]),int(row2[4]),int(row2[5]),date1,int(row1[6]),date2,int(row2[6]))
message2 = """
<p style="color:blue">美國西德州原油價格(USD/桶):</p>
<p style="color:#FF3399">%s&nbsp&nbsp 原油&nbsp%s</p>
<p style="color:#FF3399">%s&nbsp&nbsp 原油&nbsp%s</p>
<p style="color:blue">前日匯率(%s)</p>
<p style="color:#FF3399">USD/TWD&nbsp&nbsp%s</p>
<p style="color:#FF3399">CNY/TWD&nbsp&nbsp%s</p>
<p style="color:#FF3399">JPY/TWD&nbsp&nbsp%s</p>
<p style="color:#FF3399">EUR/TWD&nbsp&nbsp%s</p>
"""%(date1,row1[7],date2,row2[7],date1,row1[8],row1[9],row1[10],row1[11])
message3="""<p style="color:blue">中鋼2015年第二季7月盤價:</p>
<p style="color:blue">燁聯2015/7月份盤價:</p>
<p style="color:blue">燁聯2015/7月份盤價:/<p>
<p style="color:blue">華文專業鋼鐵網台灣地區一週鋼市(TWD/TON):</p>
<p style="color:blue">(當日匯率：USD/TWD31.70) 每週四更新/<p>
<p style="color:blue">以下附件 </p>
<p style="color:blue">LME國際金屬銅、鋁、鎳最近30天行情價格    (資料擷取台灣區電線電纜工業同業公會)</p>
<p style="color:blue"> 銅、鋁、鎳、鋅CASH BUYER價格       (資料擷取LME網站)</p>
<p style="color:blue">銅、鋁、鎳、鋅技術線圖                  (資料擷取LME網站)</p>
<p style="color:blue">美國西德州原油價格                                  (資料擷取經濟部能源局油價資訊管理與分析系統)</p>
<p style="color:blue">前日匯率                                               (資料擷取兆豐國際商業銀行)</p>
<p style="color:blue">Steelnet華文專業鋼鐵網鋼材價格                   (資料擷取華文鋼鐵鋼台灣地區一週鋼市)</p>
<p style="color:blue">請參考.</p>
<p style="color:blue">(以上資料由專案管理部提供)</p>

</b>
</body>
</html>

"""


f.write(message)
f.write(message1)
f.write(message2)
f.write(message3)
f.close()
#write into excel

from openpyxl import Workbook
from openpyxl import load_workbook
filename = 'sample.xlsx'
wb = load_workbook(filename)


# grab the active worksheet
ws = wb.active



# Rows can also be appended
exceldate = row1[1]
YDATSlice = re.split('[-]', row1[1])
print YDATSlice
EXCELInDate = ("%s/%s/%s"%(YDATSlice[0],YDATSlice[1],YDATSlice[2]))
print EXCELInDate
ws['A2']=EXCELInDate
ws['B2']=row1[2]
ws['C2']=row1[3]
ws['D2']=row1[4]
ws['E2']=row1[5]
ws['F2']=row1[6]
ws['G2']=row1[7]
ws['H2']=row1[8]
ws['I2']=row1[11]
ws['J2']=row1[9]
ws['K2']=row1[10]


# Save the file
wb.save(filename)

